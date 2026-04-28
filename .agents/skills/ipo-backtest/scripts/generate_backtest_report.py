#!/usr/bin/env python3
"""
IPO Backtest Report Generator
Generates Excel report comparing ipo-scoring scores vs actual IPO performance.

Usage:
    python generate_backtest_report.py --input ipo_data.json --scores scores.json --output backtest_report.xlsx
"""

import json
import argparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    raise


def load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_raw_data_sheet(ws, ipos: list, scores: dict):
    headers = [
        "股票代码", "公司名称", "上市日期", "发行价", "首日开盘", "首日收盘",
        "首日涨幅%", "首日破发", "累积涨跌%", "公开发售认购倍数",
        "行业", "总分", "基础分", "白/灰/黑名单", "数据来源", "备注"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, ipo in enumerate(ipos, 2):
        score = scores.get(ipo["stock_code_short"], {}) if scores else {}
        data = [
            ipo.get("stock_code", ""),
            ipo.get("company_name_cn", ""),
            ipo.get("listing_date", ""),
            ipo.get("issue_price", ""),
            ipo.get("first_day_open", ""),
            ipo.get("first_day_close", ""),
            ipo.get("first_day_return", ""),
            "是" if ipo.get("first_day_broke") else "否",
            ipo.get("cum_return", ""),
            ipo.get("public_subscription", ""),
            ipo.get("industry", ""),
            score.get("total_score", ""),
            score.get("base_score", ""),
            score.get("classification", ""),
            ipo.get("data_source", ""),
            ipo.get("notes", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 7 and isinstance(val, (int, float)):
                cell.font = Font(color="008000" if val > 0 else "CC0000", bold=True)
            if col == 14 and val:
                if "黑" in str(val):
                    cell.font = Font(color="CC0000", bold=True)
                elif "白" in str(val):
                    cell.font = Font(color="008000", bold=True)


def build_statistics_sheet(ws, ipos: list, scores: dict):
    n = len(ipos)
    broke_count = sum(1 for ipo in ipos if ipo.get("first_day_broke"))
    positive_count = sum(1 for ipo in ipos if (ipo.get("first_day_return") or 0) > 0)
    avg_return = sum(ipo.get("first_day_return", 0) for ipo in ipos) / n if n else 0
    returns = sorted([ipo.get("first_day_return", 0) for ipo in ipos])
    median_return = returns[n // 2] if n else 0
    stats = [
        ["统计指标", "数值", "说明"],
        ["样本数", n, "4月上市新股数量"],
        ["首日破发率", f"{broke_count/n*100:.1f}%" if n else "N/A", f"破发{broke_count}只"],
        ["首日正收益比例", f"{positive_count/n*100:.1f}%" if n else "N/A", f"上涨{positive_count}只"],
        ["首日平均涨幅", f"{avg_return:.2f}%", "算术平均"],
        ["首日涨幅中位数", f"{median_return:.2f}%", "中位数"],
        ["首日涨幅最高", f"{max(returns):.2f}%", ""],
        ["首日涨幅最低", f"{min(returns):.2f}%", ""],
    ]
    if scores:
        whitelist = [ipo for ipo in ipos if scores.get(ipo["stock_code_short"], {}).get("classification") == "白名单"]
        blacklist = [ipo for ipo in ipos if scores.get(ipo["stock_code_short"], {}).get("classification") == "黑名单"]
        wl_broke = sum(1 for ipo in whitelist if ipo.get("first_day_broke"))
        bl_broke = sum(1 for ipo in blacklist if ipo.get("first_day_broke"))
        stats.extend([
            ["", "", ""],
            ["评分有效性验证", "", ""],
            ["白名单数量", len(whitelist), "基础分>=80"],
            ["白名单首日破发率", f"{wl_broke/len(whitelist)*100:.1f}%" if whitelist else "N/A", "越低越好"],
            ["黑名单数量", len(blacklist), "触发黑名单条件"],
            ["黑名单首日破发率", f"{bl_broke/len(blacklist)*100:.1f}%" if blacklist else "N/A", "越高越好"],
        ])
    for row_idx, row_data in enumerate(stats, 1):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if row_idx == 1 or (len(stats) > 9 and row_idx == 10 and col == 1):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def build_group_analysis_sheet(ws, ipos: list, scores: dict):
    if not scores:
        ws.cell(row=1, column=1, value="未提供评分数据，无法生成分组分析")
        return
    high_group, mid_group, low_group = [], [], []
    for ipo in ipos:
        sc = scores.get(ipo["stock_code_short"], {})
        total = sc.get("total_score", 0)
        if total >= 80:
            high_group.append(ipo)
        elif total >= 60:
            mid_group.append(ipo)
        else:
            low_group.append(ipo)
    def group_stats(group):
        n = len(group)
        if n == 0:
            return [0, "N/A", "N/A", "N/A"]
        broke = sum(1 for ipo in group if ipo.get("first_day_broke"))
        avg = sum(ipo.get("first_day_return", 0) for ipo in group) / n
        return [n, f"{broke/n*100:.1f}%", f"{avg:.2f}%", group[0].get("company_name_cn", "") + (f"等{n}只" if n > 1 else "")]
    headers = ["分组", "数量", "首日破发率", "首日平均涨幅", "包含标的"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    groups = [
        ("高分组 (>=80分)", high_group),
        ("中分组 (60-79分)", mid_group),
        ("低分组 (<60分)", low_group),
    ]
    for row_idx, (name, group) in enumerate(groups, 2):
        stats = group_stats(group)
        ws.cell(row=row_idx, column=1, value=name)
        for col, val in enumerate(stats, 2):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 4 and isinstance(val, str) and "%" in val:
                num = float(val.replace("%", ""))
                cell.font = Font(color="008000" if num > 0 else "CC0000", bold=True)


def generate_report(ipo_data_path: str, scores_path: str, output_path: str):
    ipos = load_json(ipo_data_path)
    scores = load_json(scores_path) if scores_path else {}
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "原始数据"
    build_raw_data_sheet(ws1, ipos, scores)
    ws2 = wb.create_sheet("统计汇总")
    build_statistics_sheet(ws2, ipos, scores)
    ws3 = wb.create_sheet("分组分析")
    build_group_analysis_sheet(ws3, ipos, scores)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    wb.save(output_path)
    print(f"Backtest report saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate IPO Backtest Report")
    parser.add_argument("--input", required=True, help="Path to IPO data JSON")
    parser.add_argument("--scores", help="Path to backtest scores JSON (optional)")
    parser.add_argument("--output", required=True, help="Path to output .xlsx")
    args = parser.parse_args()
    generate_report(args.input, args.scores, args.output)


if __name__ == "__main__":
    main()
