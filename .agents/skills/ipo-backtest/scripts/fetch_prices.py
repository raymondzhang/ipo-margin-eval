#!/usr/bin/env python3
"""
Fetch real-time HK stock prices from Tencent Finance API.
No API key required. Free public endpoint.

Usage:
    python fetch_prices.py 06656 00068 03277
    python fetch_prices.py --all-tracked
    python fetch_prices.py --all-tracked --update-excel
"""

import urllib.request
import urllib.parse
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def fetch_prices(codes):
    """
    Fetch real-time prices for HK stock codes from Tencent Finance.
    codes: list of strings like ['06656', '00068']
    Returns: dict {code: {'price': float, 'name': str, 'change_pct': float, 'time': str}}
    """
    if not codes:
        return {}

    # Format: hk06656,hk00068,...
    query = ",".join(f"hk{c}" for c in codes)
    url = f"http://qt.gtimg.cn/q={query}"

    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
            }
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = resp.read()
    except Exception as e:
        print(f"Error fetching prices: {e}")
        return {}

    # Decode GBK to UTF-8
    try:
        text = data.decode("gbk", errors="replace")
    except:
        text = data.decode("utf-8", errors="replace")

    results = {}
    for line in text.strip().split(";"):
        line = line.strip()
        if not line or "=" not in line:
            continue

        # Format: v_hk06656="100~name~code~price~..."
        prefix, value = line.split("=", 1)
        code = prefix.replace('v_hk', '').replace('v_hk', '')

        # Remove quotes
        value = value.strip('"')
        parts = value.split("~")

        if len(parts) < 5:
            continue

        try:
            results[code] = {
                "name": parts[1],
                "price": float(parts[3]),
                "prev_close": float(parts[4]),
                "open": float(parts[5]) if len(parts) > 5 else None,
                "high": float(parts[33]) if len(parts) > 33 else None,
                "low": float(parts[34]) if len(parts) > 34 else None,
                "change_amt": float(parts[31]) if len(parts) > 31 else None,
                "change_pct": float(parts[32]) if len(parts) > 32 else None,
                "volume": float(parts[6]) if len(parts) > 6 else None,
                "update_time": parts[30] if len(parts) > 30 else None,
            }
        except (ValueError, IndexError) as e:
            print(f"Parse error for {code}: {e}")
            continue

    return results


def get_all_tracked_codes(excel_path):
    """Extract all HK stock codes from the tracking sheet."""
    if not openpyxl:
        print("openpyxl not installed")
        return []

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet3"]
    codes = []
    for r in range(5, ws.max_row + 1):
        val = ws.cell(row=r, column=3).value
        if val and isinstance(val, str) and val.endswith(".HK"):
            code = val.replace(".HK", "")
            codes.append(code)
    return codes


def update_excel_with_prices(excel_path, prices):
    """Update tracking Excel with fetched prices."""
    if not openpyxl:
        print("openpyxl not installed")
        return 0

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet3"]
    updated = 0

    for r in range(5, ws.max_row + 1):
        code_val = ws.cell(row=r, column=3).value
        if not code_val or not isinstance(code_val, str) or not code_val.endswith(".HK"):
            continue

        code = code_val.replace(".HK", "")
        if code not in prices:
            continue

        price_data = prices[code]
        latest_price = price_data["price"]

        # The name row is above the code row
        name_row = r - 1
        issue_price = ws.cell(row=name_row, column=7).value
        first_day_close = ws.cell(row=name_row, column=8).value

        # Update 现价 (column 12)
        ws.cell(row=name_row, column=12, value=latest_price)

        # Update 累积表现 (column 14) = (latest - issue) / issue
        if issue_price and isinstance(issue_price, (int, float)) and issue_price > 0:
            cumulative = round((latest_price - issue_price) / issue_price, 4)
            ws.cell(row=name_row, column=14, value=cumulative)

            # Color code: green if positive, red if negative
            from openpyxl.styles import Font
            if cumulative < 0:
                ws.cell(row=name_row, column=14).font = Font(color="CC0000", bold=True)
                ws.cell(row=name_row, column=12).font = Font(color="CC0000")
            else:
                ws.cell(row=name_row, column=14).font = Font(color="008000", bold=True)
                ws.cell(row=name_row, column=12).font = Font(color="008000")

        updated += 1

    wb.save(excel_path)
    return updated


def main():
    parser = argparse.ArgumentParser(description="Fetch HK stock prices")
    parser.add_argument("codes", nargs="*", help="Stock codes like 06656 00068")
    parser.add_argument("--all-tracked", action="store_true", help="Update all tracked IPOs")
    parser.add_argument("--update-excel", action="store_true", help="Write prices back to tracking Excel")
    parser.add_argument("--excel", default="/Users/rayzyp/Documents/运营流程/IPO孖展评估/IPO Score_OKK.xlsx")
    args = parser.parse_args()

    if args.all_tracked:
        codes = get_all_tracked_codes(args.excel)
        print(f"Found {len(codes)} tracked stocks: {', '.join(codes[:10])}{'...' if len(codes) > 10 else ''}")
    elif args.codes:
        codes = args.codes
    else:
        print("Usage examples:")
        print("  python fetch_prices.py 06656 00068")
        print("  python fetch_prices.py --all-tracked")
        print("  python fetch_prices.py --all-tracked --update-excel")
        return

    if not codes:
        print("No codes to fetch.")
        return

    # Fetch in batches of 50 (API limit)
    all_results = {}
    batch_size = 50
    for i in range(0, len(codes), batch_size):
        batch = codes[i:i+batch_size]
        print(f"Fetching batch {i//batch_size + 1}: {', '.join(batch[:5])}{'...' if len(batch) > 5 else ''}")
        results = fetch_prices(batch)
        all_results.update(results)

    print(f"\n{'代码':<8} {'名称':<12} {'现价':>8} {'涨跌%':>8} {'更新时间':<20}")
    print("-" * 60)
    for code in codes:
        if code in all_results:
            d = all_results[code]
            pct_str = f"{d['change_pct']:+.2f}%" if d['change_pct'] is not None else "N/A"
            print(f"{code:<8} {d['name']:<12} {d['price']:>8.2f} {pct_str:>8} {d['update_time'] or '':<20}")
        else:
            print(f"{code:<8} {'(failed)':<12}")

    if args.update_excel and args.all_tracked:
        updated = update_excel_with_prices(args.excel, all_results)
        print(f"\n✅ Updated {updated} stocks in {args.excel}")


if __name__ == "__main__":
    main()
