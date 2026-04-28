#!/usr/bin/env python3
"""
IPO Tracking Updater
Helper script to register new IPOs and update existing ones in the tracking sheet.

Usage:
    # Register a new IPO (add to tracking sheet)
    python update_tracking.py --register \
        --code 06656 --name "思格新能" --listing-date 2026-04-16 \
        --issue-price 324.2 --lot-size 100 \
        --public-subscription 1102.05 --first-day-close 659.5

    # Update existing IPOs with latest prices
    python update_tracking.py --update-prices \
        --prices '{"06656.HK": 542.5, "00068.HK": 35.2}'

    # List all tracked IPOs
    python update_tracking.py --list
"""

import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    raise


DEFAULT_TRACKING_FILE = "/Users/rayzyp/Documents/运营流程/IPO孖展评估/IPO Score_OKK.xlsx"
SHEET_NAME = "Sheet3"


def find_last_data_row(ws):
    """Find the last row that contains stock name data."""
    for r in range(ws.max_row, 4, -1):
        if ws.cell(row=r, column=3).value and isinstance(ws.cell(row=r, column=3).value, str):
            if ws.cell(row=r, column=3).value.strip() and not ws.cell(row=r, column=3).value.startswith("跌穿"):
                return r
    return ws.max_row


def find_stock_row(ws, stock_code):
    """Find the row index for a given stock code."""
    for r in range(5, ws.max_row + 1):
        val = ws.cell(row=r, column=3).value
        if val and stock_code in str(val):
            return r
    return None


def add_ipo_row(ws, stock_code, name, listing_date, issue_price, lot_size=100,
                market_cap=None, first_day_close=None, public_subscription=None,
                first_day_rate=None, cumulative_return=None):
    """Add a new IPO entry to the tracking sheet."""
    last_row = find_last_data_row(ws)
    # Find the next empty slot (entries are 2 rows: name row + code row)
    new_row = last_row + 1
    while ws.cell(row=new_row, column=3).value:
        new_row += 1

    # Row 1: Name + data
    ws.cell(row=new_row, column=3, value=name)
    ws.cell(row=new_row, column=4, value=listing_date)
    ws.cell(row=new_row, column=5, value=lot_size)
    ws.cell(row=new_row, column=6, value=market_cap)
    ws.cell(row=new_row, column=7, value=issue_price)
    ws.cell(row=new_row, column=8, value=first_day_close)
    ws.cell(row=new_row, column=9, value=public_subscription)
    ws.cell(row=new_row, column=12, value=first_day_close)  # 现价 = 首日收盘
    ws.cell(row=new_row, column=13, value=first_day_rate)   # 首日表现
    ws.cell(row=new_row, column=14, value=cumulative_return) # 累积表现

    # Row 2: Code
    code_row = new_row + 1
    ws.cell(row=code_row, column=3, value=f"{stock_code}.HK")

    print(f"Added IPO: {name} ({stock_code}.HK) at row {new_row}")
    return new_row


def update_prices(ws, prices_dict):
    """Update latest prices for tracked IPOs."""
    updated = []
    for r in range(5, ws.max_row + 1):
        code_val = ws.cell(row=r, column=3).value
        if not code_val or not isinstance(code_val, str):
            continue
        code = code_val.strip()
        if not code.endswith(".HK"):
            continue

        code_short = code.replace(".HK", "")
        if code_short not in prices_dict:
            continue

        latest_price = prices_dict[code_short]
        # The name row is above the code row
        name_row = r - 1
        issue_price = ws.cell(row=name_row, column=7).value

        if issue_price and latest_price:
            cumulative = round((latest_price - issue_price) / issue_price, 4)
            ws.cell(row=name_row, column=12, value=latest_price)   # 现价
            ws.cell(row=name_row, column=14, value=cumulative)     # 累积表现

            # Color code
            if cumulative < 0:
                ws.cell(row=name_row, column=14).font = Font(color="CC0000", bold=True)
            else:
                ws.cell(row=name_row, column=14).font = Font(color="008000", bold=True)

            updated.append({
                "code": code,
                "name": ws.cell(row=name_row, column=3).value,
                "price": latest_price,
                "cumulative": cumulative
            })

    return updated


def list_tracked_ipos(ws):
    """List all tracked IPOs."""
    ipos = []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        if not name or not isinstance(name, str):
            continue
        if name.endswith(".HK") or name.startswith("跌穿"):
            continue

        code_row = r + 1
        code = ws.cell(row=code_row, column=3).value
        listing_date = ws.cell(row=r, column=4).value
        issue_price = ws.cell(row=r, column=7).value
        latest_price = ws.cell(row=r, column=12).value
        cumulative = ws.cell(row=r, column=14).value

        ipos.append({
            "row": r,
            "name": name,
            "code": code,
            "listing_date": listing_date,
            "issue_price": issue_price,
            "latest_price": latest_price,
            "cumulative": cumulative
        })
    return ipos


def main():
    parser = argparse.ArgumentParser(description="IPO Tracking Updater")
    parser.add_argument("--file", default=DEFAULT_TRACKING_FILE, help="Path to tracking Excel file")
    parser.add_argument("--register", action="store_true", help="Register a new IPO")
    parser.add_argument("--update-prices", action="store_true", help="Update latest prices")
    parser.add_argument("--list", action="store_true", help="List all tracked IPOs")
    parser.add_argument("--code", help="Stock code (e.g., 06656)")
    parser.add_argument("--name", help="Stock name")
    parser.add_argument("--listing-date", help="Listing date (YYYY-MM-DD)")
    parser.add_argument("--issue-price", type=float, help="Issue price")
    parser.add_argument("--lot-size", type=int, default=100, help="Lot size")
    parser.add_argument("--market-cap", help="Market cap (string)")
    parser.add_argument("--first-day-close", type=float, help="First day closing price")
    parser.add_argument("--public-subscription", type=float, help="Public subscription multiple")
    parser.add_argument("--prices", help='JSON dict of prices, e.g., \'{"06656": 542.5}\'')
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without saving")

    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.file)
    ws = wb[SHEET_NAME]

    if args.register:
        if not all([args.code, args.name, args.listing_date, args.issue_price]):
            print("Error: --register requires --code, --name, --listing-date, --issue-price")
            return

        # Calculate first-day rate if first_day_close provided
        first_day_rate = None
        if args.first_day_close:
            first_day_rate = round((args.first_day_close - args.issue_price) / args.issue_price, 4)

        listing_date = datetime.strptime(args.listing_date, "%Y-%m-%d")

        if not args.dry_run:
            add_ipo_row(
                ws, args.code, args.name, listing_date, args.issue_price,
                args.lot_size, args.market_cap, args.first_day_close,
                args.public_subscription, first_day_rate, first_day_rate
            )
            wb.save(args.file)
            print(f"Saved to {args.file}")
        else:
            print(f"[DRY RUN] Would add: {args.name} ({args.code}) at {listing_date}")

    elif args.update_prices:
        if not args.prices:
            print("Error: --update-prices requires --prices '{\"06656\": 542.5}'")
            return

        prices = json.loads(args.prices)
        updated = update_prices(ws, prices)

        if not args.dry_run:
            wb.save(args.file)
            print(f"Saved to {args.file}")

        print(f"\nUpdated {len(updated)} IPOs:")
        for u in updated:
            cum_str = f"{u['cumulative']*100:+.2f}%" if u['cumulative'] else "N/A"
            print(f"  {u['code']} {u['name']}: 现价={u['price']}, 累积={cum_str}")

    elif args.list:
        ipos = list_tracked_ipos(ws)
        print(f"\nTracked IPOs ({len(ipos)} total):")
        print(f"{'Row':<5} {'Code':<10} {'Name':<15} {'Date':<12} {'Issue':<8} {'Latest':<8} {'Cum':<10}")
        print("-" * 70)
        for ipo in ipos:
            code = ipo['code'] or "N/A"
            cum_str = f"{ipo['cumulative']*100:+.2f}%" if ipo['cumulative'] else "N/A"
            date_str = ipo['listing_date'].strftime("%Y-%m-%d") if isinstance(ipo['listing_date'], datetime) else str(ipo['listing_date'])[:10]
            print(f"{ipo['row']:<5} {code:<10} {ipo['name'][:14]:<15} {date_str:<12} "
                  f"{ipo['issue_price'] or '':<8} {ipo['latest_price'] or '':<8} {cum_str:<10}")

    else:
        print("No action specified. Use --register, --update-prices, or --list.")
        parser.print_help()


if __name__ == "__main__":
    main()
